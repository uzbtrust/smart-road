"""Read a manual ASTM D6433 field survey out of the professor's spreadsheet.

Layout, one worksheet per 300 m analysis section:

    Longitudinal & Transverse Cracking
    Darajasi | Nuqson qiymatlari
    L        | 10.8 | 11 | 16.7 | ... | 334.9      <- last cell is the total
    M        | 20.1 | 10.8 | ...   | 161.6
    H        |
             |
    L density(%) | 10.336...
    M density(%) |  4.988...

Quantities are metric: square metres for area distresses, metres for linear
ones, a count for potholes. That matters, because ASTM's deduct curves were
drawn in inch-pound units and only *area* density is dimensionless -- a linear
density of 0.10 m/m2 is 0.031 ft/ft2, a factor of 3.28. Densities are therefore
recomputed here from the totals and converted before they reach the curves,
rather than taken from the sheet at face value.

The sheet's own density column is still read, and disagreements are reported:
it is a hand-built spreadsheet and some cells are wrong.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from .astm import M_PER_FT, SQ_M_PER_SQ_FT, unit_of

#: Each analysis section is 300 m of a 10.8 m carriageway (three 3.6 m lanes).
SECTION_AREA_SQ_M = 3240.0

#: Sheet heading -> ASTM distress key.
DISTRESS_ALIASES = {
    "longitudinal & transverse cracking": "longitudinal_transverse_cracking",
    "alligator cracking": "alligator_cracking",
    "block cracking": "block_cracking",
    "edge cracking": "edge_cracking",
    "slippage cracking": "slippage_cracking",
    "patching": "patching_and_utility_cut_patching",
    "potholes": "potholes",
    "bumps and sags": "bumps_and_sags",
    "shoving": "shoving",
}

SEVERITY = {"L": "low", "M": "medium", "H": "high"}


def _canonical(heading: str) -> str | None:
    h = re.sub(r"\(.*?\)", "", heading).strip().lower()
    h = re.sub(r"\s+", " ", h)
    for alias, key in DISTRESS_ALIASES.items():
        if h.startswith(alias) or alias.startswith(h):
            return key
    return None


@dataclass
class Entry:
    distress: str
    severity: str
    total: float                 # metric: m2, m, or a count
    density_pct: float           # recomputed, metric
    sheet_density_pct: float | None = None

    @property
    def density_mismatch(self) -> bool:
        if self.sheet_density_pct is None:
            return False
        return abs(self.density_pct - self.sheet_density_pct) > 0.01 * max(1.0, self.density_pct)


@dataclass
class Section:
    name: str
    area_sq_m: float
    entries: list[Entry] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


#: ASTM X1.17.1.2: a pothole wider than 750 mm is converted to an equivalent
#: number of holes by dividing its area by 0.5 m2. The survey recorded pothole
#: extent as an area (20.34, 9.79 -- not whole numbers), so that conversion is
#: applied before the count-based pothole curve is consulted.
SQ_M_PER_EQUIVALENT_POTHOLE = 0.5


def imperial_density(density_pct_metric: float, distress: str,
                     potholes_as_area: bool = True) -> float:
    """Convert a metric percent density into the units the deduct curves use.

    Area densities are ratios of like units and pass through unchanged; linear
    and count densities do not -- and a pothole density recorded as area has to
    become a count first.
    """
    unit = unit_of(distress)
    if unit == "area":
        return density_pct_metric
    if unit == "linear":
        return density_pct_metric * M_PER_FT           # m/m2 -> ft/ft2
    if potholes_as_area:
        density_pct_metric /= SQ_M_PER_EQUIVALENT_POTHOLE   # m2/m2 -> count/m2
    return density_pct_metric * SQ_M_PER_SQ_FT         # 1/m2 -> 1/ft2


def _measurements(cells) -> list[float]:
    """Numbers in a severity row, stopping at the first text cell.

    Every sheet carries a *second* block of columns further right holding the
    surveyor's own TDV/CDV/PCI figures, each introduced by a text label. Those
    columns sit on the same rows as the distress measurements, so reading a row
    to its end pulls one of them in as if it were a quantity.

    On Лист2 that is not hypothetical: the ``M`` row of Alligator Cracking is
    empty, but shares a row with ``PCI | 61``. Taking the last number invented an
    alligator-cracking-medium entry whose deduct value (27.4) was the largest in
    the section, and dropped the section's PCI by about nine points.

    Stopping at the label keeps the measurement block and its trailing sum, and
    leaves a genuinely empty severity row empty.
    """
    out: list[float] = []
    for v in cells:
        if isinstance(v, str) and v.strip():
            break
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            out.append(float(v))
    return out


#: Labels the surveyor used for his own calculation, to the right of the data.
SURVEYOR_LABELS = ("TDV", "CDV", "PCI")


def surveyor_pci(path: Path) -> dict[str, dict[str, float]]:
    """The surveyor's own TDV / CDV / PCI, read from those right-hand columns.

    The accompanying article states that a PCI was never calculated. The
    workbook says otherwise -- every sheet carries all three figures -- so they
    are available as an independent check on our own arithmetic.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[str, dict[str, float]] = {}
    for ws in wb.worksheets:
        found: dict[str, float] = {}
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().upper() in SURVEYOR_LABELS:
                    nxt = ws.cell(r, c + 1).value
                    if isinstance(nxt, (int, float)):
                        found[v.strip().upper()] = float(nxt)
        if found:
            out[ws.title] = found
    return out


def parse_workbook(path: Path, area_sq_m: float = SECTION_AREA_SQ_M) -> list[Section]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sections: list[Section] = []

    for ws in wb.worksheets:
        section = Section(name=ws.title, area_sq_m=area_sq_m)
        rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]

        current: str | None = None
        totals: dict[tuple[str, str], float] = {}
        sheet_density: dict[tuple[str, str], float] = {}

        for row in rows:
            head = str(row[0]).strip() if row[0] is not None else ""
            if not head:
                continue

            m = re.match(r"([LMH])\s*density\s*\(%\)", head, re.I)
            if m and current:
                val = next((v for v in row[1:] if isinstance(v, (int, float))), None)
                if val is not None:
                    sheet_density[(current, SEVERITY[m.group(1).upper()])] = float(val)
                continue

            if head in SEVERITY and current:
                nums = _measurements(row[1:])
                if nums:
                    # The row lists each measurement and ends with their sum.
                    totals[(current, SEVERITY[head])] = float(nums[-1])
                continue

            if head.lower().startswith("darajasi"):
                continue

            key = _canonical(head)
            if key:
                current = key
            elif head not in ("Nuqson qiymatlari",):
                section.warnings.append(f"unrecognised heading {head!r}")

        for (distress, severity), total in sorted(totals.items()):
            density = total / area_sq_m * 100.0
            section.entries.append(Entry(
                distress=distress, severity=severity, total=total,
                density_pct=density,
                sheet_density_pct=sheet_density.get((distress, severity)),
            ))
        sections.append(section)

    return sections


#: Cells whose value disagrees with the CDL published in the accompanying
#: article. In every case the article's total is recovered exactly by the
#: replacement below, so the article is right and the spreadsheet drifted.
#: (sheet, distress, severity) -> corrected metric density %
CORRECTIONS = {
    ("Лист1", "bumps_and_sags", "low"): 0.25617,                    # x10 low
    ("Лист4", "block_cracking", "high"): 3.07692,                   # quantity left undivided
    ("Лист6", "longitudinal_transverse_cracking", "low"): 5.77883,  # x10 low
    ("Лист7", "alligator_cracking", "low"): 7.16520,                # x10 low
    ("Лист7", "potholes", "high"): 0.74505,                         # density cell empty
    ("Лист8", "block_cracking", "low"): 14.68421,                   # x10 low
    ("Лист12", "longitudinal_transverse_cracking", "low"): 1.11543,  # density cell empty
}


def corrected_densities(section: Section) -> dict[tuple[str, str], float]:
    """Metric percent densities for a section, with the known bad cells replaced.

    The sheet's density column is preferred over recomputing from the totals:
    it is the column that reproduces the article's published CDL, so where the
    two disagree it is the totals that drifted.
    """
    out: dict[tuple[str, str], float] = {}
    for e in section.entries:
        key = (section.name, e.distress, e.severity)
        if key in CORRECTIONS:
            out[(e.distress, e.severity)] = CORRECTIONS[key]
        elif e.sheet_density_pct is not None:
            out[(e.distress, e.severity)] = e.sheet_density_pct
        else:
            out[(e.distress, e.severity)] = e.density_pct
    return out


def pci_from_densities(densities: dict[tuple[str, str], float]):
    """Run ASTM D6433 Sections 9.3-9.6 on metric percent densities.

    Densities are converted into the units the deduct curves were drawn in
    before they are looked up; see :func:`imperial_density`.
    """
    from .engine import (DEDUCT_FLOOR, MAX_DEDUCTS, allowable_deduct_count,
                         corrected_deduct_value, deduct_value, rate)

    deducts = sorted(
        (deduct_value(d, s, imperial_density(v, d)) for (d, s), v in densities.items() if v > 0),
        reverse=True,
    )[:MAX_DEDUCTS]
    if not deducts:
        return 100.0, "Good", 0.0, 0

    m = allowable_deduct_count(deducts[0])
    whole = int(m)
    kept = deducts[:whole]
    if m - whole > 0 and len(deducts) > whole:
        kept.append(deducts[whole] * (m - whole))

    q = sum(1 for dv in kept if dv > DEDUCT_FLOOR)
    if q <= 1:
        cdv = min(sum(kept), 100.0)
    else:
        cdv = max(
            corrected_deduct_value(sum(kept[:qi] + [min(d, DEDUCT_FLOOR) for d in kept[qi:]]), qi)
            for qi in range(1, q + 1)
        )
    pci = max(0.0, 100.0 - cdv)
    return pci, rate(pci)[0], cdv, q
